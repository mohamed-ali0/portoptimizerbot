"""
Create container type report in the exact format shown
"""
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def create_container_report():
    """
    Create the container type report in the exact format shown
    """
    print("="*60)
    print("CREATING CONTAINER TYPE REPORT")
    print("="*60)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Container Type Report"
    
    # Set up the report structure
    # Header row 1: Date
    ws['A1'] = f"Date: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Header row 2: Container Type
    ws['A2'] = "Container Type"
    ws['A2'].font = Font(bold=True, size=12)
    
    # Merge cells for container type header
    ws.merge_cells('A2:B2')
    
    # Header row 3: Column headers
    headers = [
        "Terminal", "Shipping Line", "20ST", "40ST", "40HC", "45", 
        "20RF", "40RF", "Special", "Flat"
    ]
    
    # Add column headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    # Add data rows
    row = 4
    
    # APMT Terminal data
    apmt_data = [
        ("APMT", "CMA-CGM/APL/ANL", "YES", "YES", "YES", "YES", "", "", "", ""),
        ("APMT", "Cosco", "YES", "YES", "YES", "YES", "", "", "", ""),
        ("APMT", "Evergreen", "YES", "YES", "YES", "YES", "", "", "", ""),
        ("APMT", "Hapag-Lloyd", "YES", "YES", "YES", "YES", "YES", "YES", "", ""),
        ("APMT", "Maersk", "YES", "YES", "YES", "YES", "YES", "YES", "YES", "YES"),
        ("APMT", "Hyundai", "YES", "YES", "YES", "YES", "", "", "", ""),
        ("APMT", "ONE", "YES", "YES", "YES", "YES", "", "YES", "", ""),
        ("APMT", "Yang Ming", "YES", "YES", "YES", "YES", "", "", "", ""),
        ("APMT", "ZIM", "YES", "YES", "YES", "YES", "", "", "YES", "YES"),
    ]
    
    # ITS Terminal data
    its_data = [
        ("ITS", "Hyundai", "", "", "YES", "", "", "", "", ""),
    ]
    
    # SSA-A Terminal data
    ssa_data = [
        ("SSA-A", "KMD", "YES", "YES", "YES", "", "", "", "", ""),
        ("SSA-A", "Matson", "YES", "YES", "YES", "YES", "YES", "YES", "YES", "YES"),
        ("SSA-A", "NPD", "YES", "YES", "YES", "YES", "YES", "YES", "YES", "YES"),
        ("SSA-A", "PLL", "YES", "YES", "YES", "YES", "YES", "YES", "YES", "YES"),
        ("SSA-A", "Pasha", "YES", "YES", "YES", "YES", "YES", "YES", "YES", "YES"),
        ("SSA-A", "SLS", "YES", "YES", "YES", "YES", "YES", "YES", "YES", "YES"),
        ("SSA-A", "SM Lines", "YES", "YES", "YES", "YES", "YES", "YES", "YES", "YES"),
        ("SSA-A", "TSL", "YES", "YES", "YES", "", "YES", "YES", "", ""),
    ]
    
    # Combine all data
    all_data = apmt_data + its_data + ssa_data
    
    # Add data to worksheet
    for data_row in all_data:
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='center')
        row += 1
    
    # Add borders to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply borders to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Set page orientation to landscape
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 8  # Legal size
    
    # Save the file
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = f"screenshots/container_report_{timestamp}.xlsx"
    
    wb.save(output_path)
    print(f"✅ Container report created: {output_path}")
    
    return output_path

def main():
    """Main function"""
    print("Creating container type report...")
    
    # Create the report
    excel_path = create_container_report()
    
    if excel_path:
        print(f"\n🎯 SUCCESS!")
        print(f"✅ Report created: {excel_path}")
        print("\nThe report is now ready!")
        print("You can open it in Excel and print to PDF (Ctrl+P)")
        print("It will have proper borders and landscape legal format.")
    else:
        print("\n❌ Failed to create report")

if __name__ == "__main__":
    main()
