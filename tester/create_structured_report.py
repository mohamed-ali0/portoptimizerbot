"""
Create structured report from Excel data
Converts Excel to structured JSON format and creates formatted PDF
"""
import os
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def find_latest_excel():
    """Find the most recent Excel file"""
    screenshots_dir = "screenshots"
    
    if not os.path.exists(screenshots_dir):
        return None
    
    excel_files = []
    for file in os.listdir(screenshots_dir):
        if file.endswith(('.xlsx', '.xls')):
            excel_files.append(os.path.join(screenshots_dir, file))
    
    if not excel_files:
        return None
    
    return max(excel_files, key=os.path.getctime)

def parse_excel_to_structured_data(excel_path):
    """
    Parse Excel file and convert to structured JSON format
    """
    try:
        print(f"Parsing Excel file: {excel_path}")
        
        # Load Excel file
        df = pd.read_excel(excel_path, sheet_name=0)
        print(f"Excel data shape: {df.shape}")
        print(f"Columns: {list(df.columns)}")
        
        # Create structured data
        structured_data = {
            "date": datetime.now().strftime("%Y-%m-%d"),
            "data": [],
            "key": {
                "YES": "accepting",
                "NO": "not accepting", 
                "DUAL": "dual only"
            }
        }
        
        # Process each row
        for index, row in df.iterrows():
            # Skip header rows or empty rows
            if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == "":
                continue
            
            # Extract terminal and shipping line
            terminal = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ""
            shipping_line = str(row.iloc[1]).strip() if not pd.isna(row.iloc[1]) else ""
            
            if terminal == "" or shipping_line == "":
                continue
            
            # Create container acceptance data
            container_acceptance = {}
            
            # Map container types (adjust column indices based on your Excel structure)
            container_types = ["20ST", "40ST", "40HC", "45", "20RF", "40RF", "Special", "Flat"]
            
            # Assuming the data starts from column 2 (index 2)
            for i, container_type in enumerate(container_types):
                col_index = 2 + (i * 2)  # Each container type has 2 columns (Shift 1, Shift 2)
                
                shift_1 = ""
                shift_2 = ""
                
                if col_index < len(row):
                    shift_1_val = row.iloc[col_index]
                    if not pd.isna(shift_1_val) and str(shift_1_val).strip().upper() == "YES":
                        shift_1 = "YES"
                
                if col_index + 1 < len(row):
                    shift_2_val = row.iloc[col_index + 1]
                    if not pd.isna(shift_2_val) and str(shift_2_val).strip().upper() == "YES":
                        shift_2 = "YES"
                
                container_acceptance[container_type] = {
                    "Shift 1": shift_1,
                    "Shift 2": shift_2
                }
            
            # Add to structured data
            structured_data["data"].append({
                "terminal": terminal,
                "shipping_line": shipping_line,
                "container_acceptance": container_acceptance
            })
        
        print(f"✅ Parsed {len(structured_data['data'])} shipping lines")
        return structured_data
        
    except Exception as e:
        print(f"❌ Error parsing Excel: {str(e)}")
        return None

def create_formatted_excel(structured_data, output_path):
    """
    Create a properly formatted Excel file with the structured data
    """
    try:
        print(f"Creating formatted Excel: {output_path}")
        
        # Create workbook
        wb = load_workbook()
        ws = wb.active
        ws.title = "Container Type Report"
        
        # Set up headers
        headers = [
            "Terminal", "Shipping Line", "20ST", "40ST", "40HC", "45", 
            "20RF", "40RF", "Special", "Flat"
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add data rows
        row = 2
        for item in structured_data["data"]:
            terminal = item["terminal"]
            shipping_line = item["shipping_line"]
            acceptance = item["container_acceptance"]
            
            # Create row data
            row_data = [terminal, shipping_line]
            
            # Add container acceptance data
            for container_type in ["20ST", "40ST", "40HC", "45", "20RF", "40RF", "Special", "Flat"]:
                shift_1 = acceptance[container_type]["Shift 1"]
                shift_2 = acceptance[container_type]["Shift 2"]
                
                # Combine shift data
                if shift_1 and shift_2:
                    cell_value = "YES/YES"
                elif shift_1:
                    cell_value = "YES"
                elif shift_2:
                    cell_value = "YES"
                else:
                    cell_value = ""
                
                row_data.append(cell_value)
            
            # Add row to worksheet
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal='center')
            
            row += 1
        
        # Add borders to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set page orientation to landscape
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = 8  # Legal size
        
        # Save workbook
        wb.save(output_path)
        print(f"✅ Formatted Excel created: {output_path}")
        return True
        
    except Exception as e:
        print(f"❌ Error creating formatted Excel: {str(e)}")
        return False

def save_json_data(structured_data, json_path):
    """
    Save structured data as JSON
    """
    try:
        with open(json_path, 'w') as f:
            json.dump(structured_data, f, indent=2)
        print(f"✅ JSON data saved: {json_path}")
        return True
    except Exception as e:
        print(f"❌ Error saving JSON: {str(e)}")
        return False

def main():
    """Main function"""
    print("="*60)
    print("STRUCTURED REPORT CREATOR")
    print("="*60)
    print("Converts Excel to structured format and creates formatted output")
    print("="*60)
    
    # Find latest Excel file
    excel_path = find_latest_excel()
    if not excel_path:
        print("❌ No Excel files found in screenshots directory")
        return
    
    print(f"📁 Found Excel file: {excel_path}")
    
    # Parse Excel to structured data
    structured_data = parse_excel_to_structured_data(excel_path)
    if not structured_data:
        print("❌ Failed to parse Excel file")
        return
    
    # Create timestamp for output files
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    
    # Save JSON data
    json_path = f"screenshots/structured_data_{timestamp}.json"
    save_json_data(structured_data, json_path)
    
    # Create formatted Excel
    formatted_excel_path = f"screenshots/formatted_report_{timestamp}.xlsx"
    if create_formatted_excel(structured_data, formatted_excel_path):
        print(f"\n🎯 SUCCESS!")
        print(f"✅ JSON data: {json_path}")
        print(f"✅ Formatted Excel: {formatted_excel_path}")
        print("\nThe Excel file now has proper structure and formatting!")
        print("You can now open it and print to PDF (Ctrl+P)")
    else:
        print("\n❌ Failed to create formatted Excel")

if __name__ == "__main__":
    main()
