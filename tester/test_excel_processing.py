"""
Test script for Excel processing functionality
Tests adding separator lines and converting to PDF
"""
import os
import glob
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import subprocess

def find_excel_files():
    """Find Excel files in the screenshots directory"""
    print("="*60)
    print("FINDING EXCEL FILES")
    print("="*60)
    
    # Look in screenshots directory (where they're actually being saved)
    screenshots_dir = "screenshots"
    downloads_dir = "downloads"
    
    excel_files = []
    
    # Check screenshots directory
    if os.path.exists(screenshots_dir):
        print(f"Checking {screenshots_dir} directory...")
        for file in os.listdir(screenshots_dir):
            if file.endswith(('.xlsx', '.xls')):
                excel_files.append(os.path.join(screenshots_dir, file))
                print(f"  Found: {file}")
    
    # Check downloads directory
    if os.path.exists(downloads_dir):
        print(f"Checking {downloads_dir} directory...")
        for file in os.listdir(downloads_dir):
            if file.endswith(('.xlsx', '.xls')):
                excel_files.append(os.path.join(downloads_dir, file))
                print(f"  Found: {file}")
    
    if not excel_files:
        print("❌ No Excel files found!")
        return []
    
    print(f"✅ Found {len(excel_files)} Excel file(s)")
    return excel_files

def add_separator_lines_to_excel(excel_path, output_path):
    """
    Add visible separator lines to Excel file
    """
    try:
        print(f"\n{'='*60}")
        print("ADDING SEPARATOR LINES TO EXCEL")
        print("="*60)
        print(f"Input file: {excel_path}")
        print(f"Output file: {output_path}")
        
        # Load the Excel file
        workbook = load_workbook(excel_path)
        worksheet = workbook.active
        
        print(f"Worksheet: {worksheet.title}")
        print(f"Max row: {worksheet.max_row}")
        print(f"Max column: {worksheet.max_column}")
        
        # Define border style for separator lines
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders to all cells with data
        cells_processed = 0
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
                    cells_processed += 1
        
        print(f"Processed {cells_processed} cells with data")
        
        # Save the modified Excel file
        workbook.save(output_path)
        print(f"✅ Excel file with separator lines saved: {output_path}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error adding separator lines: {str(e)}")
        return False

def convert_excel_to_pdf(excel_path, pdf_path):
    """
    Convert Excel file to PDF in landscape legal format
    """
    try:
        print(f"\n{'='*60}")
        print("CONVERTING EXCEL TO PDF")
        print("="*60)
        print(f"Input file: {excel_path}")
        print(f"Output file: {pdf_path}")
        
        # Method 1: Try LibreOffice (if available)
        try:
            print("Trying LibreOffice conversion...")
            cmd = [
                'libreoffice',
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', os.path.dirname(pdf_path),
                '--infilter=calc8',
                excel_path
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
            if result.returncode == 0:
                print("✅ LibreOffice conversion successful")
                return True
            else:
                print(f"LibreOffice failed: {result.stderr}")
        except Exception as e:
            print(f"LibreOffice not available: {str(e)}")
        
        # Method 2: Use Python libraries to create actual PDF
        try:
            print("Trying Python library conversion...")
            
            # Load Excel file
            df = pd.read_excel(excel_path, sheet_name=0)
            print(f"Excel data shape: {df.shape}")
            
            # Create a new Excel file with proper formatting first
            formatted_excel_path = pdf_path.replace('.pdf', '_formatted.xlsx')
            
            with pd.ExcelWriter(formatted_excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Sheet1', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']
                
                # Add borders to all cells
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Apply borders to all cells
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = thin_border
                
                # Set page orientation to landscape
                worksheet.page_setup.orientation = 'landscape'
                worksheet.page_setup.paperSize = 8  # Legal size
                
            print(f"✅ Formatted Excel file created: {formatted_excel_path}")
            
            # Now try to convert to PDF using different methods
            return convert_excel_to_pdf_alternative(formatted_excel_path, pdf_path)
            
        except Exception as e:
            print(f"❌ Python library conversion failed: {str(e)}")
            return False
            
    except Exception as e:
        print(f"❌ Error in PDF conversion: {str(e)}")
        return False

def convert_excel_to_pdf_alternative(excel_path, pdf_path):
    """
    Alternative PDF conversion methods
    """
    try:
        print("Trying alternative PDF conversion methods...")
        
        # Method 1: Try using weasyprint (if available)
        try:
            import weasyprint
            from weasyprint import HTML, CSS
            
            # Convert Excel to HTML first
            df = pd.read_excel(excel_path, sheet_name=0)
            
            # Create HTML table
            html_content = df.to_html(index=False, classes='table', table_id='excel-table')
            
            # Add CSS for borders and landscape layout
            css_content = """
            @page {
                size: Legal landscape;
                margin: 0.5in;
            }
            table {
                border-collapse: collapse;
                width: 100%;
                font-size: 10px;
            }
            th, td {
                border: 1px solid black;
                padding: 4px;
                text-align: left;
            }
            th {
                background-color: #f2f2f2;
                font-weight: bold;
            }
            """
            
            # Create HTML document
            html_doc = HTML(string=html_content)
            css_doc = CSS(string=css_content)
            
            # Generate PDF
            html_doc.write_pdf(pdf_path, stylesheets=[css_doc])
            print(f"✅ PDF created using weasyprint: {pdf_path}")
            return True
            
        except ImportError:
            print("weasyprint not available")
        except Exception as e:
            print(f"weasyprint failed: {str(e)}")
        
        # Method 2: Try using reportlab
        try:
            from reportlab.lib.pagesizes import legal, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            
            # Load Excel data
            df = pd.read_excel(excel_path, sheet_name=0)
            
            # Create PDF document
            doc = SimpleDocTemplate(pdf_path, pagesize=landscape(legal))
            
            # Convert DataFrame to table data
            table_data = [df.columns.tolist()] + df.values.tolist()
            
            # Create table
            table = Table(table_data)
            
            # Add table style with borders
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 6),
            ])
            
            table.setStyle(table_style)
            
            # Build PDF
            doc.build([table])
            print(f"✅ PDF created using reportlab: {pdf_path}")
            return True
            
        except ImportError:
            print("reportlab not available")
        except Exception as e:
            print(f"reportlab failed: {str(e)}")
        
        # Method 3: Try using matplotlib to create PDF
        try:
            import matplotlib.pyplot as plt
            import matplotlib.backends.backend_pdf as pdf_backend
            
            # Load Excel data
            df = pd.read_excel(excel_path, sheet_name=0)
            
            # Create figure with landscape orientation
            fig, ax = plt.subplots(figsize=(14, 8.5))  # Legal landscape size
            ax.axis('tight')
            ax.axis('off')
            
            # Create table
            table = ax.table(cellText=df.values, colLabels=df.columns, 
                           cellLoc='center', loc='center')
            
            # Style the table
            table.auto_set_font_size(False)
            table.set_fontsize(8)
            table.scale(1, 1.5)
            
            # Add borders
            for i in range(len(df.columns)):
                for j in range(len(df) + 1):
                    cell = table[(j, i)]
                    cell.set_edgecolor('black')
                    cell.set_linewidth(0.5)
            
            # Save as PDF
            plt.savefig(pdf_path, format='pdf', bbox_inches='tight', dpi=300)
            plt.close()
            print(f"✅ PDF created using matplotlib: {pdf_path}")
            return True
            
        except ImportError:
            print("matplotlib not available")
        except Exception as e:
            print(f"matplotlib failed: {str(e)}")
        
        print("❌ All PDF conversion methods failed")
        return False
        
    except Exception as e:
        print(f"❌ Error in alternative PDF conversion: {str(e)}")
        return False

def test_excel_processing():
    """Test the complete Excel processing pipeline"""
    print("="*60)
    print("EXCEL PROCESSING TEST")
    print("="*60)
    print("Testing separator lines and PDF conversion functionality")
    print("="*60)
    
    # Step 1: Find Excel files
    excel_files = find_excel_files()
    
    if not excel_files:
        print("\n❌ No Excel files found to process!")
        print("Make sure to run the Excel download first.")
        return False
    
    # Step 2: Process the most recent Excel file
    latest_file = max(excel_files, key=os.path.getctime)
    print(f"\n📁 Processing latest file: {latest_file}")
    
    # Step 3: Add separator lines
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    formatted_excel_path = f"screenshots/formatted_{timestamp}.xlsx"
    
    separator_success = add_separator_lines_to_excel(latest_file, formatted_excel_path)
    
    if not separator_success:
        print("❌ Failed to add separator lines")
        return False
    
    # Step 4: Convert to PDF
    pdf_path = f"screenshots/report_{timestamp}.pdf"
    pdf_success = convert_excel_to_pdf(formatted_excel_path, pdf_path)
    
    if not pdf_success:
        print("❌ Failed to convert to PDF")
        return False
    
    # Step 5: Summary
    print(f"\n{'='*60}")
    print("PROCESSING COMPLETE")
    print("="*60)
    print(f"✅ Original Excel: {latest_file}")
    print(f"✅ Formatted Excel: {formatted_excel_path}")
    print(f"✅ PDF Output: {pdf_path}")
    print("\n🎯 All processing steps completed successfully!")
    
    return True

def test_individual_functions():
    """Test individual functions with sample data"""
    print(f"\n{'='*60}")
    print("TESTING INDIVIDUAL FUNCTIONS")
    print("="*60)
    
    # Create a sample Excel file for testing
    sample_data = {
        'Terminal': ['APMT', 'APMT', 'ITS', 'SSA-A'],
        'Shipping Line': ['CMA-CGM', 'Maersk', 'ZIM', 'KMD'],
        '20ST': ['YES', 'YES', 'YES', 'YES'],
        '40ST': ['YES', 'YES', 'YES', 'YES'],
        '40HC': ['YES', 'YES', 'YES', 'YES'],
        '45': ['YES', 'YES', 'YES', 'YES']
    }
    
    # Create sample Excel file
    sample_excel_path = "screenshots/sample_test.xlsx"
    df = pd.DataFrame(sample_data)
    df.to_excel(sample_excel_path, index=False)
    print(f"✅ Created sample Excel file: {sample_excel_path}")
    
    # Test separator lines
    formatted_path = "screenshots/sample_formatted.xlsx"
    separator_success = add_separator_lines_to_excel(sample_excel_path, formatted_path)
    
    if separator_success:
        print("✅ Separator lines test passed")
    else:
        print("❌ Separator lines test failed")
    
    # Test PDF conversion
    pdf_path = "screenshots/sample_report.pdf"
    pdf_success = convert_excel_to_pdf(formatted_path, pdf_path)
    
    if pdf_success:
        print("✅ PDF conversion test passed")
    else:
        print("❌ PDF conversion test failed")
    
    return separator_success and pdf_success

if __name__ == "__main__":
    print("Excel Processing Test Script")
    print("="*60)
    
    # Test with actual downloaded files
    success = test_excel_processing()
    
    if not success:
        print("\n" + "="*60)
        print("TESTING WITH SAMPLE DATA")
        print("="*60)
        test_individual_functions()
    
    print("\n" + "="*60)
    print("TEST COMPLETE")
    print("="*60)
    
    input("\nPress Enter to exit...")
