"""
Convert existing Excel data to proper container report format
"""
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

def find_latest_excel():
    """Find the most recent Excel file"""
    screenshots_dir = "screenshots"
    
    if not os.path.exists(screenshots_dir):
        return None
    
    excel_files = []
    for file in os.listdir(screenshots_dir):
        if file.endswith(('.xlsx', '.xls')):
            excel_files.append(os.path.join(screenshots_dir, file))
    
    if not excel_files:
        return None
    
    return max(excel_files, key=os.path.getctime)

def convert_excel_to_proper_format(excel_path):
    """
    Convert existing Excel data to proper container report format
    """
    try:
        print(f"Converting Excel file: {excel_path}")
        
        # Load Excel file
        df = pd.read_excel(excel_path, sheet_name=0)
        print(f"Excel data shape: {df.shape}")
        print(f"Columns: {list(df.columns)}")
        
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Container Type Report"
        
        # Set up the report structure
        # Header row 1: Date
        ws['A1'] = f"Date: {datetime.now().strftime('%Y-%m-%d')}"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Header row 2: Container Type
        ws['A2'] = "Container Type"
        ws['A2'].font = Font(bold=True, size=12)
        
        # Merge cells for container type header
        ws.merge_cells('A2:B2')
        
        # Header row 3: Column headers
        headers = [
            "Terminal", "Shipping Line", "20ST", "40ST", "40HC", "45", 
            "20RF", "40RF", "Special", "Flat"
        ]
        
        # Add column headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Add data rows
        row = 4
        
        # Process each row from the original Excel
        for index, excel_row in df.iterrows():
            # Skip header rows or empty rows
            if pd.isna(excel_row.iloc[0]) or str(excel_row.iloc[0]).strip() == "":
                continue
            
            # Extract terminal and shipping line
            terminal = str(excel_row.iloc[0]).strip() if not pd.isna(excel_row.iloc[0]) else ""
            shipping_line = str(excel_row.iloc[1]).strip() if not pd.isna(excel_row.iloc[1]) else ""
            
            if terminal == "" or shipping_line == "":
                continue
            
            # Create row data
            row_data = [terminal, shipping_line]
            
            # Process container acceptance data
            # Assuming the data starts from column 2 (index 2)
            container_types = ["20ST", "40ST", "40HC", "45", "20RF", "40RF", "Special", "Flat"]
            
            for i, container_type in enumerate(container_types):
                col_index = 2 + (i * 2)  # Each container type has 2 columns (Shift 1, Shift 2)
                
                shift_1 = ""
                shift_2 = ""
                
                if col_index < len(excel_row):
                    shift_1_val = excel_row.iloc[col_index]
                    if not pd.isna(shift_1_val) and str(shift_1_val).strip().upper() == "YES":
                        shift_1 = "YES"
                
                if col_index + 1 < len(excel_row):
                    shift_2_val = excel_row.iloc[col_index + 1]
                    if not pd.isna(shift_2_val) and str(shift_2_val).strip().upper() == "YES":
                        shift_2 = "YES"
                
                # Combine shift data
                if shift_1 and shift_2:
                    cell_value = "YES"
                elif shift_1:
                    cell_value = "YES"
                elif shift_2:
                    cell_value = "YES"
                else:
                    cell_value = ""
                
                row_data.append(cell_value)
            
            # Add row to worksheet
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal='center')
            
            row += 1
        
        # Add borders to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders to all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set page orientation to landscape
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = 8  # Legal size
        
        # Save the file
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        output_path = f"screenshots/proper_format_{timestamp}.xlsx"
        
        wb.save(output_path)
        print(f"✅ Proper format Excel created: {output_path}")
        
        return output_path
        
    except Exception as e:
        print(f"❌ Error converting Excel: {str(e)}")
        return None

def main():
    """Main function"""
    print("="*60)
    print("CONVERTING EXCEL TO PROPER FORMAT")
    print("="*60)
    print("This will convert your existing Excel data to the proper format")
    print("="*60)
    
    # Find latest Excel file
    excel_path = find_latest_excel()
    if not excel_path:
        print("❌ No Excel files found in screenshots directory")
        return
    
    print(f"📁 Found Excel file: {excel_path}")
    
    # Convert to proper format
    output_path = convert_excel_to_proper_format(excel_path)
    
    if output_path:
        print(f"\n🎯 SUCCESS!")
        print(f"✅ Proper format Excel created: {output_path}")
        print("\nThe Excel file now has the proper structure!")
        print("You can open it and print to PDF (Ctrl+P)")
        print("It will have proper borders and landscape legal format.")
    else:
        print("\n❌ Failed to convert Excel")

if __name__ == "__main__":
    main()
