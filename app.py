from flask import Flask, jsonify, request, send_file, send_from_directory
from apscheduler.schedulers.background import BackgroundScheduler
from datetime import datetime, timedelta
import os
import json
import zipfile
from pathlib import Path
from automation import download_excel_report
from system_settings import SystemSettings
import threading
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import subprocess
try:
    from PyPDF2 import PdfReader, PdfWriter
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.colors import gray
    from reportlab.lib.units import inch
    PDF_LIBRARIES_AVAILABLE = True
except ImportError:
    PDF_LIBRARIES_AVAILABLE = False
    print("[WARNING] PyPDF2 or reportlab not available. Watermark functionality will be disabled.")

app = Flask(__name__)

# Initialize system settings
settings = SystemSettings()

# Base URL for publicly accessible resources
DEFAULT_BASE_DOWNLOAD_URL = "https://empties.provar.io/portoptimizer"
BASE_DOWNLOAD_URL = os.environ.get("PORTOPTIMIZER_BASE_URL") or DEFAULT_BASE_DOWNLOAD_URL
BASE_DOWNLOAD_URL = BASE_DOWNLOAD_URL.rstrip("/")


def build_download_url(filename):
    """
    Construct the public download URL for a given filename.
    """
    return f"{BASE_DOWNLOAD_URL}/download/{filename}"

def add_separator_lines_to_excel(excel_path, output_path):
    """
    Add visible separator lines to Excel file and convert to PDF
    """
    try:
        print(f"[{datetime.now()}] Adding separator lines to Excel file...")
        
        # Load the Excel file
        workbook = load_workbook(excel_path)
        worksheet = workbook.active
        
        # Define border style for separator lines
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders to all cells with data
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
        
        # Save the modified Excel file
        workbook.save(output_path)
        print(f"[{datetime.now()}] Excel file with separator lines saved: {output_path}")
        
        return True
        
    except Exception as e:
        print(f"[{datetime.now()}] Error adding separator lines: {str(e)}")
        return False

def add_watermark_to_pdf(pdf_path, watermark_text=None, watermark_data=None, use_pdf_name=False):
    """
    Add watermark to PDF file
    watermark_text: Text to display as watermark (e.g., "CONFIDENTIAL", timestamp)
    watermark_data: Dictionary of additional data to include in watermark
    use_pdf_name: If True, use PDF filename (date) as the main watermark text
    """
    if not PDF_LIBRARIES_AVAILABLE:
        print(f"[{datetime.now()}] Skipping watermark - PDF libraries not available")
        return False
    
    try:
        print(f"[{datetime.now()}] Adding watermark to PDF...")
        
        # Create watermark text - USE FULL FILENAME (without extension)
        watermark_lines = []
        
        # If use_pdf_name is True, use the full PDF filename (without extension) as watermark
        if use_pdf_name:
            pdf_filename = os.path.basename(pdf_path)
            # Remove .pdf extension to get the full name (e.g., POLA_Empty_Returns_2025-11-02_23-23-10)
            filename_without_ext = os.path.splitext(pdf_filename)[0]
            print(f"[{datetime.now()}] Watermark: PDF path='{pdf_path}', filename='{pdf_filename}', watermark_text='{filename_without_ext}'")
            
            # Ensure we have a valid filename
            if filename_without_ext:
                watermark_lines.append(filename_without_ext)
            else:
                print(f"[{datetime.now()}] WARNING: Empty filename extracted from '{pdf_path}', using fallback")
                watermark_lines.append(pdf_filename.replace('.pdf', ''))
        # Add custom text if provided (but no additional data)
        elif watermark_text:
            watermark_lines.append(watermark_text)
        
        # If no watermark specified, use default with timestamp
        if not watermark_lines:
            timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
            watermark_lines.append(f"Generated: {timestamp}")
        
        watermark_text_final = "\n".join(watermark_lines)
        
        # Read the original PDF
        reader = PdfReader(pdf_path)
        writer = PdfWriter()
        
        # Get page dimensions (Legal landscape: 14 x 8.5 inches)
        # Legal size: 14" x 8.5" (landscape)
        page_width = 14 * inch
        page_height = 8.5 * inch
        
        # Create watermark overlay with Legal landscape dimensions
        watermark_path = pdf_path.replace('.pdf', '_watermark_temp.pdf')
        c = canvas.Canvas(watermark_path, pagesize=(page_width, page_height))
        
        # Set watermark properties (light gray, semi-transparent, horizontal)
        # Increase font size by 80%: 24 * 1.8 = 43.2, round to 43
        font_size = int(24 * 1.8)  # 43
        c.setFillColor(gray, alpha=0.3)  # 30% opacity
        c.setFont("Helvetica-Bold", font_size)
        
        # Position at top of page, 2 pixels from top (horizontal, not rotated)
        # PDF coordinates: origin (0,0) is at bottom-left, top is at page_height
        # 2 pixels = 2 points (ReportLab uses points as units)
        top_offset = 2  # 2 pixels from top
        
        # Split text into lines for multiline support
        watermark_lines_list = watermark_text_final.split('\n')
        line_height = font_size * 1.2  # Line spacing (120% of font size)
        
        # Draw each line of watermark text (horizontal, top-aligned)
        # Position: 2 pixels from top, centered horizontally
        for i, line in enumerate(watermark_lines_list):
            if line.strip():  # Only draw non-empty lines
                text_width = c.stringWidth(line, "Helvetica-Bold", font_size)
                # X position: centered horizontally
                x_pos = (page_width - text_width) / 2
                # Y position: top of page minus 2 pixels, then down for each line
                # drawString uses baseline, so subtract font_size to position top of text at offset
                y_pos = page_height - top_offset - font_size - (i * line_height)
                c.drawString(x_pos, y_pos, line)
        
        c.save()
        
        # Read watermark PDF
        watermark_reader = PdfReader(watermark_path)
        watermark_page = watermark_reader.pages[0]
        
        # Overlay watermark on all pages
        for page in reader.pages:
            # Merge watermark with page
            page.merge_page(watermark_page)
            writer.add_page(page)
        
        # Write watermarked PDF
        with open(pdf_path, 'wb') as output_file:
            writer.write(output_file)
        
        # Clean up temporary watermark file
        if os.path.exists(watermark_path):
            os.remove(watermark_path)
        
        print(f"[{datetime.now()}] Watermark added successfully")
        return True
        
    except Exception as e:
        print(f"[{datetime.now()}] Error adding watermark: {str(e)}")
        # If watermark fails, return True anyway (PDF exists without watermark)
        return True

def convert_excel_to_pdf(excel_path, pdf_path, watermark_text=None, watermark_data=None, use_pdf_name=False):
    """
    Convert Excel file to PDF with borders in landscape legal format using PowerShell
    Then add watermark if specified
    watermark_text: Text to display as watermark (e.g., "CONFIDENTIAL", timestamp)
    watermark_data: Dictionary of additional data to include in watermark
    use_pdf_name: If True, use PDF filename (date) as the main watermark text
    """
    try:
        print(f"[{datetime.now()}] Converting Excel to PDF with borders...")
        
        # Create PowerShell script for Excel to PDF conversion with borders
        ps_script = f'''
        $excel = New-Object -ComObject Excel.Application
        $excel.Visible = $false
        $excel.DisplayAlerts = $false
        
        try {{
            $workbook = $excel.Workbooks.Open('{os.path.abspath(excel_path)}')
            $worksheet = $workbook.Worksheets.Item(1)
            
            # Add borders to all cells with data
            $usedRange = $worksheet.UsedRange
            if ($usedRange) {{
                $usedRange.Borders.LineStyle = 1  # xlContinuous
                $usedRange.Borders.Weight = 1     # xlThin
                # Set gray color for all border sides
                $usedRange.Borders.Item(7).Color = 12632256  # xlEdgeLeft - Light gray
                $usedRange.Borders.Item(8).Color = 12632256  # xlEdgeTop - Light gray  
                $usedRange.Borders.Item(9).Color = 12632256  # xlEdgeBottom - Light gray
                $usedRange.Borders.Item(10).Color = 12632256 # xlEdgeRight - Light gray
            }}
            
            # Set to Legal landscape
            $worksheet.PageSetup.Orientation = 2  # xlLandscape
            $worksheet.PageSetup.PaperSize = 5     # xlPaperLegal
            
            # Export to PDF
            $pdf_path = '{os.path.abspath(pdf_path)}'
            $workbook.ExportAsFixedFormat(0, $pdf_path)
            
            $workbook.Close()
            $excel.Quit()
            
            Write-Host "PDF created: $pdf_path"
        }} catch {{
            Write-Host "Error: $_"
            $excel.Quit()
        }}
        '''
        
        # Write PowerShell script to file
        with open('temp.ps1', 'w') as f:
            f.write(ps_script)
        
        try:
            # Run PowerShell script
            result = subprocess.run(['powershell', '-ExecutionPolicy', 'Bypass', '-File', 'temp.ps1'], 
                                  capture_output=True, text=True, timeout=60)
            
            if result.returncode == 0:
                print(f"[{datetime.now()}] PDF conversion successful: {pdf_path}")
                
                # Add watermark if specified
                if watermark_text or watermark_data or use_pdf_name:
                    add_watermark_to_pdf(pdf_path, watermark_text, watermark_data, use_pdf_name)
                
                return True
            else:
                print(f"[{datetime.now()}] PowerShell error: {result.stderr}")
                return False
                
        except subprocess.TimeoutExpired:
            print(f"[{datetime.now()}] PDF conversion timed out")
            return False
        except Exception as e:
            print(f"[{datetime.now()}] Error: {e}")
            return False
        finally:
            # Clean up
            if os.path.exists('temp.ps1'):
                os.remove('temp.ps1')
            
    except Exception as e:
        print(f"[{datetime.now()}] Error in PDF conversion: {str(e)}")
        return False

# Initialize screenshots directory
SCREENSHOTS_DIR = "screenshots"
os.makedirs(SCREENSHOTS_DIR, exist_ok=True)

# Scheduler for automated screenshots
scheduler = BackgroundScheduler()
screenshot_lock = threading.Lock()


def scheduled_excel_download_task():
    """Task to download Excel report - runs according to frequency"""
    with screenshot_lock:
        try:
            print(f"[{datetime.now()}] Downloading scheduled Excel report...")
            credentials = settings.get_login_credentials()
            success, message = download_excel_report(
                credentials['username'], 
                credentials['password']
            )
            if success:
                print(f"[{datetime.now()}] Excel download completed successfully: {message}")
                
                # Convert Excel to PDF with borders
                excel_path = os.path.join("downloads", message)
                if os.path.exists(excel_path):
                    # Create pdfs subdirectory
                    pdfs_dir = os.path.join("downloads", "pdfs")
                    os.makedirs(pdfs_dir, exist_ok=True)
                    
                    # Convert directly to PDF with borders
                    pdf_filename = message.replace('.xlsx', '.pdf')
                    pdf_path = os.path.join(pdfs_dir, pdf_filename)
                    
                    # Get watermark settings
                    watermark_settings = settings.get_watermark_settings()
                    watermark_text = None
                    watermark_data = None  # Only date, no additional data
                    use_pdf_name = watermark_settings.get("use_pdf_name", True)  # Default to True
                    
                    print(f"[{datetime.now()}] Watermark settings: enabled={watermark_settings.get('enabled')}, use_pdf_name={use_pdf_name}, pdf_filename={pdf_filename}")
                    
                    if watermark_settings.get("enabled", True):
                        # Only use PDF name - no timestamp or additional data
                        if use_pdf_name:
                            # Use full PDF filename (will be extracted in watermark function)
                            pass
                        elif watermark_settings.get("text"):
                            watermark_text = watermark_settings["text"]
                    
                    if convert_excel_to_pdf(excel_path, pdf_path, watermark_text, watermark_data, use_pdf_name):
                        print(f"[{datetime.now()}] PDF created successfully: {pdf_path}")
                    else:
                        print(f"[{datetime.now()}] PDF conversion failed")
            else:
                print(f"[{datetime.now()}] Excel download failed: {message}")
                # Schedule retry in 5 minutes
                print(f"[{datetime.now()}] Scheduling retry in 5 minutes...")
                scheduler.add_job(
                    func=scheduled_excel_download_task,
                    trigger="date",
                    run_date=datetime.now() + timedelta(minutes=5),
                    id='screenshot_retry',
                    replace_existing=True
                )
        except Exception as e:
            print(f"[{datetime.now()}] Error in scheduled screenshot: {str(e)}")
            # Schedule retry in 5 minutes on exception
            print(f"[{datetime.now()}] Scheduling retry in 5 minutes after exception...")
            try:
                scheduler.add_job(
                    func=scheduled_excel_download_task,
                    trigger="date",
                    run_date=datetime.now() + timedelta(minutes=5),
                    id='screenshot_retry',
                    replace_existing=True
                )
            except:
                pass


def restart_scheduler():
    """Restart the scheduler with updated frequency"""
    scheduler.remove_all_jobs()
    frequency_hours = settings.get_frequency()
    scheduler.add_job(
        func=scheduled_excel_download_task,
        trigger="interval",
        hours=frequency_hours,
        id='screenshot_job',
        replace_existing=True
    )
    print(f"Scheduler restarted with frequency: {frequency_hours} hours")


@app.route('/')
def index():
    """API information endpoint"""
    return jsonify({
        "name": "PortOptimizer Screenshot API",
        "version": "1.0.0",
        "endpoints": {
            "GET /screenshot/<date>": "Get screenshot by date (format: YYYY-MM-DD)",
            "GET /screenshots/range": "Get screenshots range (params: start_date, end_date or last_n)",
            "POST /admin/frequency": "Change screenshot frequency (requires admin_password, frequency_hours)",
            "POST /admin/credentials": "Update login credentials (requires admin_password, username, password)",
            "POST /admin/cleanup": "Delete all screenshots (requires admin_password)",
            "POST /screenshot/now": "Take screenshot immediately (requires admin_password)"
        }
    })


@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    """
    Download a specific file (Excel or PDF)
    """
    try:
        # Check downloads/pdfs directory first
        pdfs_dir = os.path.join("downloads", "pdfs")
        if os.path.exists(pdfs_dir):
            file_path = os.path.join(pdfs_dir, filename)
            if os.path.exists(file_path):
                return send_from_directory(
                    pdfs_dir, 
                    filename,
                    as_attachment=True,
                    download_name=filename
                )
        
        # Check downloads directory
        downloads_dir = "downloads"
        if os.path.exists(downloads_dir):
            file_path = os.path.join(downloads_dir, filename)
            if os.path.exists(file_path):
                return send_from_directory(
                    downloads_dir, 
                    filename,
                    as_attachment=True,
                    download_name=filename
                )
        
        # Check screenshots directory as fallback
        if os.path.exists(SCREENSHOTS_DIR):
            file_path = os.path.join(SCREENSHOTS_DIR, filename)
            if os.path.exists(file_path):
                return send_from_directory(
                    SCREENSHOTS_DIR, 
                    filename,
                    as_attachment=True,
                    download_name=filename
                )
        
        return jsonify({
            "success": False,
            "error": "File not found"
        }), 404
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/excel/<date_str>', methods=['GET'])
def get_excel_report(date_str):
    """
    Get Excel report for a specific date
    Date format: YYYY-MM-DD
    Returns JSON with download link
    """
    try:
        # Validate date format
        target_date = datetime.strptime(date_str, "%Y-%m-%d")
        
        # Find Excel files for that date
        # Files now start with "POLA_Empty_Returns_" prefix
        date_pattern = target_date.strftime("%Y-%m-%d")
        matching_files = []
        
        # Check downloads directory
        downloads_dir = "downloads"
        if os.path.exists(downloads_dir):
            for filename in os.listdir(downloads_dir):
                # Check if filename contains the date pattern and ends with .xlsx/.xls
                # Format: POLA_Empty_Returns_YYYY-MM-DD_HH-MM-SS.xlsx
                if f"POLA_Empty_Returns_{date_pattern}" in filename and filename.endswith(('.xlsx', '.xls')):
                    matching_files.append(filename)
        
        if not matching_files:
            return jsonify({
                "success": False,
                "error": f"No Excel report found for date {date_str}"
            }), 404
        
        # Return the most recent Excel file for that date
        matching_files.sort(reverse=True)
        excel_file = matching_files[0]
        
        # Return JSON with download link
        download_url = build_download_url(excel_file)
        
        return jsonify({
            "success": True,
            "date": date_str,
            "filename": excel_file,
            "download_url": download_url,
            "message": f"Excel report found for {date_str}"
        })
    
    except ValueError:
        return jsonify({
            "success": False,
            "error": "Invalid date format. Use YYYY-MM-DD"
        }), 400
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/pdf/<date_str>', methods=['GET'])
def get_pdf_report(date_str):
    """
    Get PDF report for a specific date
    Date format: YYYY-MM-DD
    Returns JSON with download link
    """
    try:
        # Validate date format
        target_date = datetime.strptime(date_str, "%Y-%m-%d")
        
        # Find PDF files for that date
        # Files now start with "POLA_Empty_Returns_" prefix
        date_pattern = target_date.strftime("%Y-%m-%d")
        matching_files = []
        
        # Check downloads/pdfs directory
        pdfs_dir = os.path.join("downloads", "pdfs")
        if os.path.exists(pdfs_dir):
            for filename in os.listdir(pdfs_dir):
                # Format: POLA_Empty_Returns_YYYY-MM-DD_HH-MM-SS.pdf
                if f"POLA_Empty_Returns_{date_pattern}" in filename and filename.endswith('.pdf'):
                    matching_files.append(filename)
        
        # Check downloads directory as fallback
        downloads_dir = "downloads"
        if os.path.exists(downloads_dir):
            for filename in os.listdir(downloads_dir):
                # Format: POLA_Empty_Returns_YYYY-MM-DD_HH-MM-SS.pdf
                if f"POLA_Empty_Returns_{date_pattern}" in filename and filename.endswith('.pdf'):
                    matching_files.append(filename)
        
        if not matching_files:
            return jsonify({
                "success": False,
                "error": f"No PDF report found for date {date_str}"
            }), 404
        
        # Return the most recent PDF file for that date
        matching_files.sort(reverse=True)
        pdf_file = matching_files[0]
        
        # Return JSON with download link
        download_url = build_download_url(pdf_file)
        
        return jsonify({
            "success": True,
            "date": date_str,
            "filename": pdf_file,
            "download_url": download_url,
            "message": f"PDF report found for {date_str}"
        })
    
    except ValueError:
        return jsonify({
            "success": False,
            "error": "Invalid date format. Use YYYY-MM-DD"
        }), 400
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/screenshots/range', methods=['GET'])
def get_screenshots_range():
    """
    Get screenshots for a date range or last N screenshots
    Parameters:
    - start_date: YYYY-MM-DD (required if not using last_n)
    - end_date: YYYY-MM-DD (required if not using last_n)
    - last_n: number of most recent screenshots (optional)
    Returns JSON with download link for ZIP file
    """
    try:
        last_n = request.args.get('last_n', type=int)
        
        if last_n:
            # Get last N screenshots
            all_files = sorted(
                [f for f in os.listdir(SCREENSHOTS_DIR) if f.endswith('.png')],
                reverse=True
            )
            selected_files = all_files[:last_n]
        else:
            # Get screenshots in date range
            start_date_str = request.args.get('start_date')
            end_date_str = request.args.get('end_date')
            
            if not start_date_str or not end_date_str:
                return jsonify({
                    "success": False,
                    "error": "Provide either 'last_n' or both 'start_date' and 'end_date'"
                }), 400
            
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
            
            if start_date > end_date:
                return jsonify({
                    "success": False,
                    "error": "start_date must be before or equal to end_date"
                }), 400
            
            # Find all screenshots in range
            selected_files = []
            for filename in os.listdir(SCREENSHOTS_DIR):
                if filename.endswith('.png'):
                    try:
                        # Extract date from filename (format: YYYY-MM-DD_HH-MM-SS.png)
                        file_date_str = filename.split('_')[0]
                        file_date = datetime.strptime(file_date_str, "%Y-%m-%d")
                        
                        if start_date <= file_date <= end_date:
                            selected_files.append(filename)
                    except:
                        continue
            
            selected_files.sort()
        
        if not selected_files:
            return jsonify({
                "success": False,
                "error": "No screenshots found for the specified criteria"
            }), 404
        
        # Create zip file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        zip_filename = f"screenshots_{timestamp}.zip"
        zip_path = os.path.join(SCREENSHOTS_DIR, zip_filename)
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for filename in selected_files:
                file_path = os.path.join(SCREENSHOTS_DIR, filename)
                zipf.write(file_path, filename)
        
        # Return JSON with download link
        download_url = build_download_url(zip_filename)
        
        return jsonify({
            "success": True,
            "zip_filename": zip_filename,
            "download_url": download_url,
            "screenshot_count": len(selected_files),
            "screenshots": selected_files,
            "message": f"ZIP file created with {len(selected_files)} screenshot(s)"
        })
    
    except ValueError as e:
        return jsonify({
            "success": False,
            "error": "Invalid date format. Use YYYY-MM-DD"
        }), 400
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/admin/frequency', methods=['POST'])
def change_frequency():
    """
    Change screenshot frequency
    Body: {
        "admin_password": "password",
        "frequency_hours": 24
    }
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "error": "Request body must be JSON"
            }), 400
        
        admin_password = data.get('admin_password')
        frequency_hours = data.get('frequency_hours')
        
        if not admin_password or frequency_hours is None:
            return jsonify({
                "success": False,
                "error": "admin_password and frequency_hours are required"
            }), 400
        
        # Verify admin password
        if not settings.verify_admin_password(admin_password):
            return jsonify({
                "success": False,
                "error": "Invalid admin password"
            }), 403
        
        # Validate frequency
        if not isinstance(frequency_hours, (int, float)) or frequency_hours <= 0:
            return jsonify({
                "success": False,
                "error": "frequency_hours must be a positive number"
            }), 400
        
        # Update frequency
        settings.set_frequency(frequency_hours)
        restart_scheduler()
        
        return jsonify({
            "success": True,
            "message": f"Frequency updated to {frequency_hours} hours",
            "frequency_hours": frequency_hours
        })
    
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/admin/credentials', methods=['POST'])
def update_credentials():
    """
    Update login credentials for the portal
    Body: {
        "admin_password": "password",
        "username": "new_username",
        "password": "new_password"
    }
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "error": "Request body must be JSON"
            }), 400
        
        admin_password = data.get('admin_password')
        username = data.get('username')
        password = data.get('password')
        
        if not admin_password or not username or not password:
            return jsonify({
                "success": False,
                "error": "admin_password, username, and password are required"
            }), 400
        
        # Verify admin password
        if not settings.verify_admin_password(admin_password):
            return jsonify({
                "success": False,
                "error": "Invalid admin password"
            }), 403
        
        # Update credentials
        settings.set_login_credentials(username, password)
        
        return jsonify({
            "success": True,
            "message": "Login credentials updated successfully"
        })
    
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/admin/cleanup', methods=['POST'])
def cleanup():
    """
    Delete all screenshots
    Body: {
        "admin_password": "password"
    }
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "error": "Request body must be JSON"
            }), 400
        
        admin_password = data.get('admin_password')
        
        if not admin_password:
            return jsonify({
                "success": False,
                "error": "admin_password is required"
            }), 400
        
        # Verify admin password
        if not settings.verify_admin_password(admin_password):
            return jsonify({
                "success": False,
                "error": "Invalid admin password"
            }), 403
        
        # Delete all screenshots
        deleted_count = 0
        for filename in os.listdir(SCREENSHOTS_DIR):
            if filename.endswith('.png') or filename.endswith('.zip'):
                file_path = os.path.join(SCREENSHOTS_DIR, filename)
                try:
                    os.remove(file_path)
                    deleted_count += 1
                except Exception as e:
                    print(f"Error deleting {filename}: {str(e)}")
        
        return jsonify({
            "success": True,
            "message": f"Cleanup completed. {deleted_count} files deleted."
        })
    
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/admin/preferred_hour', methods=['POST'])
def set_preferred_hour():
    """
    Set preferred hour for scheduled captures
    Body: {
        "admin_password": "password",
        "preferred_hour": 10
    }
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "error": "Request body must be JSON"
            }), 400
        
        admin_password = data.get('admin_password')
        preferred_hour = data.get('preferred_hour')
        
        if not admin_password or preferred_hour is None:
            return jsonify({
                "success": False,
                "error": "admin_password and preferred_hour are required"
            }), 400
        
        # Verify admin password
        if not settings.verify_admin_password(admin_password):
            return jsonify({
                "success": False,
                "error": "Invalid admin password"
            }), 403
        
        # Validate hour
        if not isinstance(preferred_hour, int) or preferred_hour < 0 or preferred_hour > 23:
            return jsonify({
                "success": False,
                "error": "preferred_hour must be an integer between 0 and 23"
            }), 400
        
        # Update preferred hour
        settings.set_preferred_hour(preferred_hour)
        
        return jsonify({
            "success": True,
            "message": f"Preferred hour updated to {preferred_hour}:00",
            "preferred_hour": preferred_hour
        })
    
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/admin/watermark', methods=['POST'])
def set_watermark_settings():
    """
    Configure watermark settings
    Body: {
        "admin_password": "password",
        "enabled": true/false (optional),
        "text": "CONFIDENTIAL" (optional, null to remove),
        "use_pdf_name": true/false (optional, use PDF filename/date as watermark),
        "include_timestamp": true/false (optional),
        "include_date": true/false (optional)
    }
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "error": "Request body must be JSON"
            }), 400
        
        admin_password = data.get('admin_password')
        
        if not admin_password:
            return jsonify({
                "success": False,
                "error": "admin_password is required"
            }), 400
        
        # Verify admin password
        if not settings.verify_admin_password(admin_password):
            return jsonify({
                "success": False,
                "error": "Invalid admin password"
            }), 403
        
        # Update watermark settings
        settings.set_watermark_settings(
            enabled=data.get('enabled'),
            text=data.get('text'),
            use_pdf_name=data.get('use_pdf_name'),
            include_timestamp=data.get('include_timestamp'),
            include_date=data.get('include_date')
        )
        
        # Get updated settings
        watermark_settings = settings.get_watermark_settings()
        
        return jsonify({
            "success": True,
            "message": "Watermark settings updated",
            "watermark": watermark_settings
        })
    
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/excel/now', methods=['POST'])
def download_excel_now():
    """
    Download Excel report immediately
    Body: {
        "admin_password": "password"
    }
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "error": "Request body must be JSON"
            }), 400
        
        admin_password = data.get('admin_password')
        
        if not admin_password:
            return jsonify({
                "success": False,
                "error": "admin_password is required"
            }), 400
        
        # Verify admin password
        if not settings.verify_admin_password(admin_password):
            return jsonify({
                "success": False,
                "error": "Invalid admin password"
            }), 403
        
        # Download Excel report
        with screenshot_lock:
            credentials = settings.get_login_credentials()
            success, message = download_excel_report(
                credentials['username'], 
                credentials['password']
            )
        
        if success:
            # Convert Excel to PDF with borders
            excel_path = os.path.join("downloads", message)
            if os.path.exists(excel_path):
                # Create pdfs subdirectory
                pdfs_dir = os.path.join("downloads", "pdfs")
                os.makedirs(pdfs_dir, exist_ok=True)
                
                # Convert directly to PDF with borders
                pdf_filename = message.replace('.xlsx', '.pdf')
                pdf_path = os.path.join(pdfs_dir, pdf_filename)
                
                # Get watermark settings
                watermark_settings = settings.get_watermark_settings()
                watermark_text = None
                watermark_data = None  # Only date, no additional data
                use_pdf_name = watermark_settings.get("use_pdf_name", True)  # Default to True
                
                print(f"[{datetime.now()}] Watermark settings: enabled={watermark_settings.get('enabled')}, use_pdf_name={use_pdf_name}, pdf_filename={pdf_filename}")
                
                if watermark_settings.get("enabled", True):
                    # Only use PDF name - no timestamp or additional data
                    if use_pdf_name:
                        # Use full PDF filename (will be extracted in watermark function)
                        pass
                    elif watermark_settings.get("text"):
                        watermark_text = watermark_settings["text"]
                
                if convert_excel_to_pdf(excel_path, pdf_path, watermark_text, watermark_data, use_pdf_name):
                    print(f"[{datetime.now()}] PDF created successfully: {pdf_path}")
                    return jsonify({
                        "success": True,
                        "message": "Excel report downloaded and PDF created successfully",
                        "excel_filename": message,
                        "pdf_filename": pdf_filename,
                        "download_url": build_download_url(pdf_filename)
                    })
                else:
                    print(f"[{datetime.now()}] PDF conversion failed")
                    return jsonify({
                        "success": True,
                        "message": "Excel report downloaded but PDF conversion failed",
                        "excel_filename": message,
                        "download_url": build_download_url(message)
                    })
            else:
                return jsonify({
                    "success": True,
                    "message": "Excel report downloaded successfully",
                    "filename": message,
                    "download_url": build_download_url(message)
                })
        else:
            return jsonify({
                "success": False,
                "error": message
            }), 500
    
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/status', methods=['GET'])
def status():
    """Get current system status"""
    try:
        frequency = settings.get_frequency()
        preferred_hour = settings.get_preferred_hour()
        credentials = settings.get_login_credentials()
        
        # Count files
        screenshot_count = len([f for f in os.listdir(SCREENSHOTS_DIR) if f.endswith('.png')])
        
        # Count Excel and PDF files
        downloads_dir = "downloads"
        pdfs_dir = os.path.join("downloads", "pdfs")
        
        excel_count = 0
        pdf_count = 0
        
        if os.path.exists(downloads_dir):
            excel_count = len([f for f in os.listdir(downloads_dir) if f.endswith(('.xlsx', '.xls'))])
        
        if os.path.exists(pdfs_dir):
            pdf_count = len([f for f in os.listdir(pdfs_dir) if f.endswith('.pdf')])
        
        # Get last screenshot time
        screenshots = sorted(
            [f for f in os.listdir(SCREENSHOTS_DIR) if f.endswith('.png')],
            reverse=True
        )
        last_screenshot = screenshots[0] if screenshots else None
        
        # Get watermark settings
        watermark_settings = settings.get_watermark_settings()
        
        return jsonify({
            "success": True,
            "frequency_hours": frequency,
            "preferred_hour": preferred_hour,
            "username": credentials['username'],
            "watermark": watermark_settings,
            "total_screenshots": screenshot_count,
            "total_excel_files": excel_count,
            "total_pdf_files": pdf_count,
            "last_screenshot": last_screenshot,
            "scheduler_running": scheduler.running
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


if __name__ == '__main__':
    # Start the scheduler
    scheduler.start()
    restart_scheduler()
    
    print("=" * 50)
    print("PortOptimizer Screenshot API")
    print("=" * 50)
    print(f"Screenshot frequency: {settings.get_frequency()} hours")
    print(f"Screenshots directory: {SCREENSHOTS_DIR}")
    print("Scheduler started")
    print("=" * 50)
    
    try:
        app.run(host='0.0.0.0', port=5004, debug=False)
    except (KeyboardInterrupt, SystemExit):
        scheduler.shutdown()
        print("\nShutdown complete")

